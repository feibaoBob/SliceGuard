"""Calculate Line Hit Rate (LHR) for misuse detection results.

LHR measures line-level localization accuracy:
- For each true positive file (correctly detected misuse), if any reported
  line number intersects with the ground truth misuse line numbers, it counts
  as a line hit.
- LHR = (line hits) / (true positive files)
"""

import csv
import os
from openpyxl import load_workbook


def parse_line_numbers(raw) -> set[int]:
    """Parse a tab/comma/newline-separated line number string into a set of ints."""
    if raw is None:
        return set()
    text = str(raw).strip().replace("\t", "").replace("\n", "").strip()
    if not text or text.lower() == "none":
        return set()
    return {int(x.strip()) for x in text.split(",") if x.strip().isdigit()}


def extract_filename_key(file_path: str) -> str:
    """Extract the unique identifying part of the file path for matching."""
    return os.path.basename(file_path)


def load_ground_truth(xlsx_path: str) -> dict[str, set[int]]:
    """Load ground truth from Excel, returning {filename: {line_numbers}}."""
    wb = load_workbook(xlsx_path)
    ws = wb.active
    gt = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        file_path, line_nums = row[0], row[1]
        if file_path is None:
            continue
        key = extract_filename_key(str(file_path))
        lines = parse_line_numbers(line_nums)
        if lines:
            gt[key] = lines
    wb.close()
    return gt


def _detect_encoding(csv_path: str) -> str:
    """Detect file encoding by trying to decode the full content."""
    with open(csv_path, "rb") as f:
        raw = f.read()
    for enc in ["utf-8-sig", "utf-8", "gbk", "gb18030", "latin1"]:
        try:
            raw.decode(enc)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "latin1"


def load_analysis_results(csv_path: str) -> list[dict]:
    """Load analysis results from CSV."""
    results = []
    enc = _detect_encoding(csv_path)
    with open(csv_path, "r", encoding=enc) as f:
        reader = csv.DictReader(f)
        for row in reader:
            file_path = row["file_path"]
            key = extract_filename_key(file_path)
            reported_lines = parse_line_numbers(row["Misuse Line Numbers"])
            rule_number = row["Rule Numbers"].strip().replace("\t", "")
            results.append({
                "file_path": file_path,
                "key": key,
                "reported_lines": reported_lines,
                "detected": rule_number != "-1" and len(reported_lines) > 0,
            })
    return results


def calculate_lhr(ground_truth: dict[str, set[int]], analysis_results: list[dict]):
    """Calculate Line Hit Rate."""
    true_positive_files = 0
    line_hits = 0
    details = []

    for result in analysis_results:
        key = result["key"]
        reported = result["reported_lines"]
        detected = result["detected"]

        gt_lines = ground_truth.get(key)

        if not detected:
            details.append((key, reported, gt_lines, "NOT_DETECTED", set()))
            continue

        if gt_lines is None:
            details.append((key, reported, None, "NOT_IN_GT", set()))
            continue

        true_positive_files += 1
        intersection = reported & gt_lines
        hit = len(intersection) > 0
        if hit:
            line_hits += 1

        status = "LINE_HIT" if hit else "LINE_MISS"
        details.append((key, reported, gt_lines, status, intersection))

    lhr = line_hits / true_positive_files if true_positive_files > 0 else 0.0
    return lhr, true_positive_files, line_hits, details


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    gt_path = os.path.join(base_dir, "Ground-Truth.xlsx")
    # ar_path = os.path.join(base_dir, "analysis_result.csv")
    # ar_path = os.path.join(base_dir, "analysis_result_by_Pro_moonshotai_Kimi-K2.6-withCoT_134327.csv")
    # ar_path = os.path.join(base_dir, "analysis_result_by_Pro_zai-org_GLM-5.1-withCoT_132050.csv")
    ar_path = os.path.join(base_dir, "analysis_result_by_qwen3-coder-next-withCoT_101701.csv")
    # ar_path = os.path.join(base_dir, "analysis_result_by_qwen3-next-80b-a3b-instruct-withCoT_091301.csv")

    ground_truth = load_ground_truth(gt_path)
    analysis_results = load_analysis_results(ar_path)

    lhr, tp_count, hit_count, details = calculate_lhr(ground_truth, analysis_results)

    print(f"{'File':<55} {'Reported':<15} {'Ground Truth':<15} {'Status':<12} {'Intersection'}")
    print("-" * 120)
    for key, reported, gt_lines, status, intersection in details:
        rep_str = str(sorted(reported)) if reported else "{}"
        gt_str = str(sorted(gt_lines)) if gt_lines else "{}"
        inter_str = str(sorted(intersection)) if intersection else ""
        print(f"{key:<55} {rep_str:<15} {gt_str:<15} {status:<12} {inter_str}")

    print(f"\n{'='*60}")
    print(f"True Positive Files: {tp_count}")
    print(f"Line Hits:           {hit_count}")
    print(f"Line Hit Rate (LHR): {lhr:.4f} ({lhr*100:.2f}%)")


if __name__ == "__main__":
    main()
